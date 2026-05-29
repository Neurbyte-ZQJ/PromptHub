from fastapi import FastAPI, Depends, HTTPException, Query, Response, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import asc, desc, or_, func as sa_func
from typing import List, Optional
from io import BytesIO
from openpyxl import Workbook, load_workbook
from datetime import timedelta
from passlib.context import CryptContext

from .database import get_db, engine, Base
from .models import Prompt, Tag, PromptVersion, User, Favorite, Category, SharedLink, PromptCollaborator, prompt_tag_association, prompt_category_association
from .schemas import (
    PromptCreate,
    PromptUpdate,
    PromptListItem,
    PromptResponse,
    TagResponse,
    CategoryCreate,
    CategoryUpdate,
    CategoryResponse,
    UserCreate,
    LoginRequest,
    UserResponse,
    Token,
    RefreshRequest,
    PasswordResetRequest,
    PasswordReset,
    FavoriteResponse,
    SharedLinkCreate,
    SharedLinkResponse,
    SharedPromptResponse,
    SharedLinkAccess,
    CollaboratorAdd,
    CollaboratorUpdate,
    CollaboratorResponse,
)
from .auth import (
    verify_password,
    get_password_hash,
    create_access_token,
    create_refresh_token,
    verify_refresh_token,
    create_reset_token,
    verify_reset_token,
    get_current_user,
)

Base.metadata.create_all(bind=engine)

share_pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")

app = FastAPI(
    title="PromptHub API",
    description="AI 提示词资产管理库",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Total-Count", "X-Page", "X-Page-Size", "X-Total-Pages"],
)


def attach_favorite_info(prompt: Prompt, user_id: int, db: Session):
    prompt.favorite_count = db.query(sa_func.count(Favorite.id)).filter(Favorite.prompt_id == prompt.id).scalar() or 0
    prompt.is_favorited = db.query(Favorite).filter(Favorite.prompt_id == prompt.id, Favorite.user_id == user_id).first() is not None


def attach_collaborator_role(prompt: Prompt, user_id: int, db: Session):
    if prompt.user_id == user_id:
        prompt.collaborator_role = "owner"
    else:
        collab = db.query(PromptCollaborator).filter(
            PromptCollaborator.prompt_id == prompt.id,
            PromptCollaborator.user_id == user_id,
        ).first()
        prompt.collaborator_role = collab.role if collab else None


def build_category_tree(categories: List[Category]) -> List[Category]:
    """将扁平分类列表构建为树形结构"""
    cat_map = {c.id: c for c in categories}
    roots = []
    for c in categories:
        if c.parent_id is None:
            roots.append(c)
        elif c.parent_id in cat_map:
            parent = cat_map[c.parent_id]
            if c not in parent.children:
                parent.children.append(c)
    return roots


@app.post("/api/auth/register", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def register(user_data: UserCreate, db: Session = Depends(get_db)):
    if db.query(User).filter(User.username == user_data.username).first():
        raise HTTPException(status_code=400, detail="用户名已存在")
    if db.query(User).filter(User.email == user_data.email).first():
        raise HTTPException(status_code=400, detail="邮箱已被注册")

    db_user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=get_password_hash(user_data.password),
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user


@app.post("/api/auth/login", response_model=Token)
def login(user_data: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == user_data.email).first()
    if not user or not verify_password(user_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="邮箱或密码错误",
            headers={"WWW-Authenticate": "Bearer"},
        )

    access_token = create_access_token(data={"sub": str(user.id)})
    refresh_token = create_refresh_token(data={"sub": str(user.id)})
    return Token(access_token=access_token, refresh_token=refresh_token, token_type="bearer")


@app.post("/api/auth/refresh", response_model=Token)
def refresh_token(request: RefreshRequest, db: Session = Depends(get_db)):
    user_id = verify_refresh_token(request.refresh_token)
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="用户不存在",
        )
    new_access_token = create_access_token(data={"sub": str(user.id)})
    new_refresh_token = create_refresh_token(data={"sub": str(user.id)})
    return Token(access_token=new_access_token, refresh_token=new_refresh_token, token_type="bearer")


@app.get("/api/auth/me", response_model=UserResponse)
def read_current_user(current_user: User = Depends(get_current_user)):
    return current_user


@app.post("/api/auth/password-reset-request")
def request_password_reset(request_data: PasswordResetRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == request_data.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="该邮箱未注册")

    reset_token = create_reset_token(data={"sub": str(user.id)})
    return {"reset_token": reset_token, "message": "重置令牌已生成"}


@app.post("/api/auth/password-reset")
def reset_password(reset_data: PasswordReset, db: Session = Depends(get_db)):
    user_id = verify_reset_token(reset_data.token)
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    user.hashed_password = get_password_hash(reset_data.new_password)
    db.commit()
    return {"message": "密码重置成功"}


@app.post("/api/prompts", response_model=PromptResponse)
def create_prompt(
    prompt_data: PromptCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tag_ids = list(prompt_data.tag_ids)
    category_ids = list(prompt_data.category_ids)
    prompt_dict = prompt_data.model_dump(exclude={"tag_ids", "new_tags", "category_ids"})

    for tag_name in prompt_data.new_tags:
        existing = db.query(Tag).filter(Tag.name == tag_name).first()
        if existing:
            if existing.id not in tag_ids:
                tag_ids.append(existing.id)
        else:
            new_tag = Tag(name=tag_name)
            db.add(new_tag)
            db.flush()
            tag_ids.append(new_tag.id)

    db_prompt = Prompt(**prompt_dict, user_id=current_user.id)

    if tag_ids:
        tags = db.query(Tag).filter(Tag.id.in_(tag_ids)).all()
        db_prompt.tags = tags

    if category_ids:
        categories = db.query(Category).filter(Category.id.in_(category_ids)).all()
        db_prompt.categories = categories

    db.add(db_prompt)
    db.flush()

    first_version = PromptVersion(
        prompt_id=db_prompt.id,
        content=db_prompt.content,
        version_number=1,
    )
    db.add(first_version)
    db.commit()
    db.refresh(db_prompt)
    attach_favorite_info(db_prompt, current_user.id, db)
    attach_collaborator_role(db_prompt, current_user.id, db)
    return db_prompt


@app.get("/api/prompts", response_model=List[PromptListItem])
def list_prompts(
    response: Response,
    search: Optional[str] = Query(None, description="按标题或内容搜索"),
    tag_id: Optional[int] = Query(None, description="按标签ID筛选"),
    category_id: Optional[int] = Query(None, description="按分类ID筛选"),
    favorites_only: bool = Query(False, description="仅返回收藏的提示词"),
    page: Optional[int] = Query(None, ge=1, description="页码，从1开始"),
    page_size: Optional[int] = Query(None, ge=1, le=100, description="每页数量，最大100"),
    skip: Optional[int] = Query(None, ge=0, description="跳过的记录数"),
    limit: Optional[int] = Query(None, ge=1, le=100, description="返回的记录数，最大100"),
    sort_by: str = Query("created_at", description="排序字段：created_at、updated_at、title、id"),
    sort_order: str = Query("desc", description="排序方向：asc、desc"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Prompt).filter(
        or_(
            Prompt.user_id == current_user.id,
            Prompt.is_public == True,
            Prompt.id.in_(db.query(PromptCollaborator.prompt_id).filter(PromptCollaborator.user_id == current_user.id)),
        )
    )

    if search:
        pattern = f"%{search}%"
        query = query.filter(
            (Prompt.title.ilike(pattern)) | (Prompt.content.ilike(pattern))
        )

    if tag_id is not None:
        query = query.join(prompt_tag_association).filter(
            prompt_tag_association.c.tag_id == tag_id
        )

    if category_id is not None:
        # 筛选当前分类及其所有子分类下的提示词（多对多）
        def get_descendant_ids(cat_id: int) -> List[int]:
            ids = [cat_id]
            children = db.query(Category).filter(Category.parent_id == cat_id).all()
            for child in children:
                ids.extend(get_descendant_ids(child.id))
            return ids
        all_cat_ids = get_descendant_ids(category_id)
        query = query.join(prompt_category_association).filter(
            prompt_category_association.c.category_id.in_(all_cat_ids)
        )

    if favorites_only:
        favorited_prompt_ids = db.query(Favorite.prompt_id).filter(Favorite.user_id == current_user.id)
        query = query.filter(Prompt.id.in_(favorited_prompt_ids))

    sort_columns = {
        "created_at": Prompt.created_at,
        "updated_at": Prompt.updated_at,
        "title": Prompt.title,
        "id": Prompt.id,
    }
    if sort_by not in sort_columns:
        raise HTTPException(status_code=400, detail="不支持的排序字段")
    if sort_order not in {"asc", "desc"}:
        raise HTTPException(status_code=400, detail="不支持的排序方向")

    total = query.count()
    sort_column = sort_columns[sort_by]
    query = query.order_by(asc(sort_column) if sort_order == "asc" else desc(sort_column), Prompt.id.desc())

    offset = skip
    effective_limit = limit
    current_page = page
    current_page_size = page_size

    if page is not None or page_size is not None:
        current_page = page or 1
        current_page_size = page_size or 20
        offset = (current_page - 1) * current_page_size
        effective_limit = current_page_size
    elif limit is not None:
        offset = skip or 0
        current_page_size = limit
        current_page = (offset // limit) + 1

    response.headers["X-Total-Count"] = str(total)
    if current_page is not None:
        response.headers["X-Page"] = str(current_page)
    if current_page_size is not None:
        response.headers["X-Page-Size"] = str(current_page_size)
        response.headers["X-Total-Pages"] = str((total + current_page_size - 1) // current_page_size)

    if offset is not None:
        query = query.offset(offset)
    if effective_limit is not None:
        query = query.limit(effective_limit)

    prompts = query.all()
    for p in prompts:
        attach_favorite_info(p, current_user.id, db)
        attach_collaborator_role(p, current_user.id, db)
    return prompts


# ============ 批量导入/导出 API ============

@app.get("/api/prompts/import-template")
def download_import_template():
    """下载导入模板 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "提示词导入模板"

    # 表头
    headers = ["标题", "场景", "内容", "变量", "是否公开", "分类", "标签"]
    ws.append(headers)

    # 示例数据
    ws.append([
        "代码审查助手",
        "代码质量检查",
        "请对以下代码进行审查，找出潜在的问题和改进建议：\n\n{{code}}",
        "code:需要审查的代码",
        "否",
        "开发工具",
        "编程,代码审查,助手"
    ])
    ws.append([
        "周报生成器",
        "工作汇报",
        "请根据以下工作内容，生成一份结构清晰的周报：\n\n{{tasks}}",
        "tasks:本周完成的任务",
        "是",
        "工作效率",
        "写作,汇报"
    ])

    # 说明 sheet
    ws2 = wb.create_sheet(title="填写说明")
    instructions = [
        ["字段", "是否必填", "说明", "示例"],
        ["标题", "必填", "提示词的名称，简短概括用途", "代码审查助手"],
        ["场景", "选填", "提示词适用的场景描述", "代码质量检查"],
        ["内容", "必填", "提示词的正文内容，支持用 {{变量名}} 标记变量", "请对以下代码进行审查：{{code}}"],
        ["变量", "选填", '变量定义，格式为"变量名:说明"，多个变量用逗号分隔', "code:需要审查的代码"],
        ["是否公开", "选填", '是否对其他用户可见，填写"是"或"否"，默认为"否"', "否"],
        ["分类", "选填", '所属分类名称，多个分类用英文逗号分隔，不存在的分类会自动创建', "开发工具,工作效率"],
        ["标签", "选填", "标签名称，多个标签用英文逗号分隔，不存在的标签会自动创建", "编程,代码审查,助手"],
    ]
    for row in instructions:
        ws2.append(row)

    # 设置列宽
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
            ws_sheet.column_dimensions[col_letter].width = max(max_length + 4, 12)

    # 说明 sheet 的列宽单独调整
    ws2.column_dimensions["C"].width = 55
    ws2.column_dimensions["D"].width = 35

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prompts_import_template.xlsx"},
    )


@app.get("/api/prompts/export")
def export_prompts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出当前用户的所有提示词为 Excel"""
    prompts = db.query(Prompt).filter(Prompt.user_id == current_user.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "提示词"
    ws.append(["标题", "场景", "内容", "变量", "是否公开", "分类", "标签"])

    for p in prompts:
        ws.append([
            p.title,
            p.scenario or "",
            p.content,
            p.variables or "",
            "是" if p.is_public else "否",
            ",".join(c.name for c in p.categories),
            ",".join(t.name for t in p.tags),
        ])

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prompts_export.xlsx"},
    )


@app.post("/api/prompts/import")
def import_prompts(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从 Excel 文件批量导入提示词"""
    try:
        content = file.file.read()
        wb = load_workbook(filename=BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="无效的 Excel 文件")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    imported = 0
    skipped = 0
    for row in rows:
        if not row or not row[0] or not row[2]:
            skipped += 1
            continue

        title = str(row[0]).strip()
        scenario = str(row[1]).strip() if row[1] else None
        content_val = str(row[2]).strip()
        variables = str(row[3]).strip() if row[3] else None
        is_public = str(row[4]).strip() == "是" if row[4] else False
        categories_str = str(row[5]).strip() if row[5] else ""
        tags_str = str(row[6]).strip() if row[6] else ""

        if not title or not content_val:
            skipped += 1
            continue

        # 处理分类（多分类，逗号分隔）
        category_ids = []
        if categories_str:
            for cat_name in categories_str.split(","):
                cat_name = cat_name.strip()
                if not cat_name:
                    continue
                cat = db.query(Category).filter(
                    Category.name == cat_name,
                    or_(Category.user_id == current_user.id, Category.user_id == None)
                ).first()
                if cat:
                    if cat.id not in category_ids:
                        category_ids.append(cat.id)
                else:
                    new_cat = Category(name=cat_name, user_id=current_user.id)
                    db.add(new_cat)
                    db.flush()
                    category_ids.append(new_cat.id)

        tag_ids = []
        if tags_str:
            for tag_name in tags_str.split(","):
                tag_name = tag_name.strip()
                if not tag_name:
                    continue
                existing = db.query(Tag).filter(Tag.name == tag_name).first()
                if existing:
                    if existing.id not in tag_ids:
                        tag_ids.append(existing.id)
                else:
                    new_tag = Tag(name=tag_name)
                    db.add(new_tag)
                    db.flush()
                    tag_ids.append(new_tag.id)

        db_prompt = Prompt(
            title=title,
            scenario=scenario,
            content=content_val,
            variables=variables,
            is_public=is_public,
            user_id=current_user.id,
        )
        if tag_ids:
            tags = db.query(Tag).filter(Tag.id.in_(tag_ids)).all()
            db_prompt.tags = tags
        if category_ids:
            cats = db.query(Category).filter(Category.id.in_(category_ids)).all()
            db_prompt.categories = cats

        db.add(db_prompt)
        db.flush()

        first_version = PromptVersion(
            prompt_id=db_prompt.id,
            content=db_prompt.content,
            version_number=1,
        )
        db.add(first_version)
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped}


@app.get("/api/prompts/{prompt_id}", response_model=PromptResponse)
def get_prompt(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    is_owner = prompt.user_id == current_user.id
    is_collaborator = db.query(PromptCollaborator).filter(
        PromptCollaborator.prompt_id == prompt_id,
        PromptCollaborator.user_id == current_user.id,
    ).first() is not None
    if not is_owner and not prompt.is_public and not is_collaborator:
        raise HTTPException(status_code=403, detail="无权访问该提示词")
    attach_favorite_info(prompt, current_user.id, db)
    attach_collaborator_role(prompt, current_user.id, db)
    return prompt


@app.put("/api/prompts/{prompt_id}", response_model=PromptResponse)
def update_prompt(
    prompt_id: int,
    prompt_data: PromptUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    is_owner = prompt.user_id == current_user.id
    collab = db.query(PromptCollaborator).filter(
        PromptCollaborator.prompt_id == prompt_id,
        PromptCollaborator.user_id == current_user.id,
    ).first()
    is_editor = collab and collab.role == "editor"
    if not is_owner and not is_editor:
        raise HTTPException(status_code=403, detail="无权修改该提示词")

    update_dict = prompt_data.model_dump(exclude_unset=True)
    tag_ids = update_dict.pop("tag_ids", None)
    new_tags = update_dict.pop("new_tags", [])
    category_ids = update_dict.pop("category_ids", None)

    content_changed = False
    if "content" in update_dict and update_dict["content"] != prompt.content:
        content_changed = True

    for key, value in update_dict.items():
        setattr(prompt, key, value)

    if category_ids is not None:
        cats = db.query(Category).filter(Category.id.in_(category_ids)).all()
        prompt.categories = cats

    if tag_ids is not None or new_tags:
        final_tag_ids = list(tag_ids) if tag_ids is not None else [t.id for t in prompt.tags]
        for tag_name in new_tags:
            existing = db.query(Tag).filter(Tag.name == tag_name).first()
            if existing:
                if existing.id not in final_tag_ids:
                    final_tag_ids.append(existing.id)
            else:
                new_tag = Tag(name=tag_name)
                db.add(new_tag)
                db.flush()
                final_tag_ids.append(new_tag.id)
        tags = db.query(Tag).filter(Tag.id.in_(final_tag_ids)).all()
        prompt.tags = tags

    if content_changed:
        latest_version = (
            db.query(PromptVersion)
            .filter(PromptVersion.prompt_id == prompt.id)
            .order_by(PromptVersion.version_number.desc())
            .first()
        )
        next_version = (latest_version.version_number + 1) if latest_version else 1
        new_version = PromptVersion(
            prompt_id=prompt.id,
            content=update_dict["content"],
            version_number=next_version,
        )
        db.add(new_version)

    db.commit()
    db.refresh(prompt)
    attach_favorite_info(prompt, current_user.id, db)
    attach_collaborator_role(prompt, current_user.id, db)
    return prompt


@app.delete("/api/prompts/{prompt_id}")
def delete_prompt(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    if prompt.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="无权删除该提示词")

    db.delete(prompt)
    db.commit()
    return {"message": "删除成功"}


@app.get("/api/tags", response_model=List[TagResponse])
def list_tags(db: Session = Depends(get_db)):
    return db.query(Tag).order_by(Tag.name).all()


@app.post("/api/prompts/{prompt_id}/favorite", response_model=FavoriteResponse)
def toggle_favorite(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    if prompt.user_id != current_user.id and not prompt.is_public:
        raise HTTPException(status_code=403, detail="无权操作该提示词")

    existing = db.query(Favorite).filter(
        Favorite.user_id == current_user.id,
        Favorite.prompt_id == prompt_id,
    ).first()

    if existing:
        db.delete(existing)
        db.commit()
        return existing
    else:
        favorite = Favorite(user_id=current_user.id, prompt_id=prompt_id)
        db.add(favorite)
        db.commit()
        db.refresh(favorite)
        return favorite


# ============ 分类 CRUD API ============

@app.get("/api/categories", response_model=List[CategoryResponse])
def list_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # 1. 当前用户自己的分类 + 公共分类
    own_categories = db.query(Category).filter(
        or_(Category.user_id == current_user.id, Category.user_id == None)
    ).all()

    # 2. 公开提示词中引用的、属于其他用户的分类（外部分类）
    visible_prompt_ids = db.query(Prompt.id).filter(
        or_(Prompt.user_id == current_user.id, Prompt.is_public == True)
    ).subquery()
    foreign_cat_ids = db.query(prompt_category_association.c.category_id).filter(
        prompt_category_association.c.prompt_id.in_(visible_prompt_ids)
    ).distinct().all()
    foreign_cat_ids_set = {r[0] for r in foreign_cat_ids}

    if foreign_cat_ids_set:
        foreign_categories = db.query(Category).filter(
            Category.id.in_(foreign_cat_ids_set),
            Category.user_id != None,
            Category.user_id != current_user.id,
        ).all()
        # 标记为外部分类，附加 owner_username
        for fc in foreign_categories:
            fc.is_foreign = True
            owner = db.query(User).filter(User.id == fc.user_id).first()
            fc.owner_username = owner.username if owner else None
    else:
        foreign_categories = []

    # 合并去重
    own_ids = {c.id for c in own_categories}
    all_categories = list(own_categories)
    for fc in foreign_categories:
        if fc.id not in own_ids:
            all_categories.append(fc)

    all_categories.sort(key=lambda c: (c.sort_order, c.name))

    # 清空 children 避免重复，手动构建树
    for c in all_categories:
        c.children = []
    tree = build_category_tree(all_categories)
    return tree


@app.post("/api/categories", response_model=CategoryResponse, status_code=status.HTTP_201_CREATED)
def create_category(
    category_data: CategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if category_data.parent_id is not None:
        parent = db.query(Category).filter(Category.id == category_data.parent_id).first()
        if not parent:
            raise HTTPException(status_code=400, detail="父分类不存在")

    db_category = Category(
        name=category_data.name,
        parent_id=category_data.parent_id,
        sort_order=category_data.sort_order,
        user_id=current_user.id,
    )
    db.add(db_category)
    db.commit()
    db.refresh(db_category)
    db_category.children = []
    return db_category


@app.put("/api/categories/{category_id}", response_model=CategoryResponse)
def update_category(
    category_id: int,
    category_data: CategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    category = db.query(Category).filter(Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类未找到")
    if category.user_id is not None and category.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="无权修改该分类")

    update_dict = category_data.model_dump(exclude_unset=True)

    # 防止将分类设为自己的子分类（循环引用）
    if "parent_id" in update_dict and update_dict["parent_id"] is not None:
        if update_dict["parent_id"] == category_id:
            raise HTTPException(status_code=400, detail="不能将分类设为自己的子分类")
        # 检查目标父分类是否是当前分类的后代
        def is_descendant(parent_id: int, child_id: int) -> bool:
            children = db.query(Category).filter(Category.parent_id == parent_id).all()
            for child in children:
                if child.id == child_id or is_descendant(child.id, child_id):
                    return True
            return False
        if is_descendant(category_id, update_dict["parent_id"]):
            raise HTTPException(status_code=400, detail="不能将分类移动到自己的子分类下")

    for key, value in update_dict.items():
        setattr(category, key, value)

    db.commit()
    db.refresh(category)
    category.children = []
    return category


@app.delete("/api/categories/{category_id}")
def delete_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    category = db.query(Category).filter(Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="分类未找到")
    if category.user_id is not None and category.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="无权删除该分类")

    # 子分类的 parent_id 设为被删分类的 parent_id（提升一级）
    for child in db.query(Category).filter(Category.parent_id == category_id).all():
        child.parent_id = category.parent_id

    # 多对多关联表会自动清除关联关系，无需手动处理提示词

    db.delete(category)
    db.commit()
    return {"message": "删除成功"}


# ============ 分享链接 API ============

@app.post("/api/prompts/{prompt_id}/shares", response_model=SharedLinkResponse, status_code=status.HTTP_201_CREATED)
def create_share_link(
    prompt_id: int,
    share_data: SharedLinkCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    if prompt.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="仅所有者可创建分享链接")

    hashed_password = None
    if share_data.password:
        hashed_password = share_pwd_ctx.hash(share_data.password)

    expires_at = None
    if share_data.expires_hours:
        from datetime import datetime as dt
        expires_at = dt.utcnow() + timedelta(hours=share_data.expires_hours)

    shared_link = SharedLink(
        prompt_id=prompt_id,
        token=SharedLink.generate_token(),
        password=hashed_password,
        expires_at=expires_at,
        created_by=current_user.id,
    )
    db.add(shared_link)
    db.commit()
    db.refresh(shared_link)
    shared_link.has_password = bool(hashed_password)
    return shared_link


@app.get("/api/prompts/{prompt_id}/shares", response_model=List[SharedLinkResponse])
def list_share_links(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    if prompt.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="仅所有者可查看分享链接")

    links = db.query(SharedLink).filter(SharedLink.prompt_id == prompt_id).all()
    for link in links:
        link.has_password = bool(link.password)
    return links


@app.delete("/api/prompts/{prompt_id}/shares/{share_id}")
def delete_share_link(
    prompt_id: int,
    share_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    link = db.query(SharedLink).filter(SharedLink.id == share_id, SharedLink.prompt_id == prompt_id).first()
    if not link:
        raise HTTPException(status_code=404, detail="分享链接未找到")
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt or prompt.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="仅所有者可删除分享链接")

    db.delete(link)
    db.commit()
    return {"message": "分享链接已删除"}


@app.get("/api/shared/{token}", response_model=SharedPromptResponse)
def access_shared_prompt(
    token: str,
    password: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    link = db.query(SharedLink).filter(SharedLink.token == token).first()
    if not link:
        raise HTTPException(status_code=404, detail="分享链接无效")

    from datetime import datetime as dt
    if link.expires_at and link.expires_at < dt.utcnow():
        raise HTTPException(status_code=410, detail="分享链接已过期")

    if link.password:
        if not password or not share_pwd_ctx.verify(password, link.password):
            raise HTTPException(status_code=403, detail="访问密码错误")

    prompt = db.query(Prompt).filter(Prompt.id == link.prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")

    return SharedPromptResponse(
        title=prompt.title,
        scenario=prompt.scenario,
        content=prompt.content,
        variables=prompt.variables,
        owner_username=prompt.owner_username,
        categories=[CategoryResponse.model_validate(c, from_attributes=True) for c in prompt.categories],
        tags=[TagResponse.model_validate(t, from_attributes=True) for t in prompt.tags],
    )


# ============ 协作者 API ============

@app.post("/api/prompts/{prompt_id}/collaborators", response_model=CollaboratorResponse, status_code=status.HTTP_201_CREATED)
def add_collaborator(
    prompt_id: int,
    collab_data: CollaboratorAdd,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    if prompt.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="仅所有者可添加协作者")

    if collab_data.role not in ("viewer", "editor"):
        raise HTTPException(status_code=400, detail="角色必须是 viewer 或 editor")

    target_user = db.query(User).filter(User.id == collab_data.user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="用户未找到")

    if target_user.id == prompt.user_id:
        raise HTTPException(status_code=400, detail="不能将所有者添加为协作者")

    existing = db.query(PromptCollaborator).filter(
        PromptCollaborator.prompt_id == prompt_id,
        PromptCollaborator.user_id == collab_data.user_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该用户已是协作者")

    collaborator = PromptCollaborator(
        prompt_id=prompt_id,
        user_id=collab_data.user_id,
        role=collab_data.role,
    )
    db.add(collaborator)
    db.commit()
    db.refresh(collaborator)
    collaborator.username = target_user.username
    return collaborator


@app.get("/api/prompts/{prompt_id}/collaborators", response_model=List[CollaboratorResponse])
def list_collaborators(
    prompt_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    if prompt.user_id != current_user.id:
        is_collab = db.query(PromptCollaborator).filter(
            PromptCollaborator.prompt_id == prompt_id,
            PromptCollaborator.user_id == current_user.id,
        ).first()
        if not is_collab:
            raise HTTPException(status_code=403, detail="无权查看协作者列表")

    collaborators = db.query(PromptCollaborator).filter(PromptCollaborator.prompt_id == prompt_id).all()
    for c in collaborators:
        c.username = c.user.username if c.user else None
    return collaborators


@app.put("/api/prompts/{prompt_id}/collaborators/{user_id}", response_model=CollaboratorResponse)
def update_collaborator(
    prompt_id: int,
    user_id: int,
    collab_data: CollaboratorUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    if prompt.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="仅所有者可修改协作者角色")

    if collab_data.role not in ("viewer", "editor"):
        raise HTTPException(status_code=400, detail="角色必须是 viewer 或 editor")

    collaborator = db.query(PromptCollaborator).filter(
        PromptCollaborator.prompt_id == prompt_id,
        PromptCollaborator.user_id == user_id,
    ).first()
    if not collaborator:
        raise HTTPException(status_code=404, detail="协作者未找到")

    collaborator.role = collab_data.role
    db.commit()
    db.refresh(collaborator)
    collaborator.username = collaborator.user.username if collaborator.user else None
    return collaborator


@app.delete("/api/prompts/{prompt_id}/collaborators/{user_id}")
def remove_collaborator(
    prompt_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    prompt = db.query(Prompt).filter(Prompt.id == prompt_id).first()
    if not prompt:
        raise HTTPException(status_code=404, detail="提示词未找到")
    is_owner = prompt.user_id == current_user.id
    is_self = user_id == current_user.id
    if not is_owner and not is_self:
        raise HTTPException(status_code=403, detail="无权移除协作者")

    collaborator = db.query(PromptCollaborator).filter(
        PromptCollaborator.prompt_id == prompt_id,
        PromptCollaborator.user_id == user_id,
    ).first()
    if not collaborator:
        raise HTTPException(status_code=404, detail="协作者未找到")

    db.delete(collaborator)
    db.commit()
    return {"message": "协作者已移除"}


@app.get("/api/users/search", response_model=List[UserResponse])
def search_users(
    q: str = Query(..., min_length=1, description="搜索关键词"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    users = db.query(User).filter(
        or_(
            User.username.ilike(f"%{q}%"),
            User.email.ilike(f"%{q}%"),
        ),
        User.id != current_user.id,
    ).limit(10).all()
    return users

